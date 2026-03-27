"""XLSX export with openpyxl — 4 formatted sheets."""

from __future__ import annotations

import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from .models.glossary import GlossaryConfig
from .models.translation import BatchResult
from .models.units import TranslationUnit
from .models.validation import ValidationReport, ValidationSeverity

# -- Helpers -----------------------------------------------------------------

# Same regex openpyxl uses internally — strip illegal XML 1.0 control chars
_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _clean(text: str | None) -> str | None:
    """Remove control characters that openpyxl rejects."""
    if text is None:
        return None
    return _ILLEGAL_XML_RE.sub("", text)


# -- Styles ------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, size=12, color="FFFFFF", name="Arial")
_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_SECTION_FONT = Font(bold=True, size=11, color="2F5496", name="Arial")
_SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
_NORMAL_FONT = Font(size=10, name="Arial")
_NOTE_FONT = Font(size=9, italic=True, color="666666", name="Arial")
_ERROR_FONT = Font(size=10, color="CC0000", name="Arial")
_WARNING_FONT = Font(size=10, color="CC7700", name="Arial")
_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _apply_header_row(ws, row: int, values: list[str]) -> None:
    """Write and style a header row."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
        cell.border = _BORDER


def _apply_cell_style(ws, row: int, col: int, font: Font = _NORMAL_FONT) -> None:
    cell = ws.cell(row=row, column=col)
    cell.font = font
    cell.alignment = _WRAP
    cell.border = _BORDER


def _write_translation_sheet(
    wb: Workbook,
    batches: list[BatchResult],
    pages: list[list[TranslationUnit]],
    target_lang: str,
) -> None:
    """Main translation sheet."""
    ws = wb.active
    ws.title = "Tlumaczenie"

    headers = [
        "Strona",
        "Sekcja / Kategoria",
        "Tekst oryginalny (PL/EN)",
        f"Tlumaczenie ({target_lang.upper()})",
        "Uwagi dla grafika",
    ]
    _apply_header_row(ws, 1, headers)

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 25

    ws.freeze_panes = "A2"

    row = 2
    last_section = ""

    for batch in batches:
        # Section header row
        if batch.section_name and batch.section_name != last_section:
            last_section = batch.section_name
            section_label = f"STRONA {batch.page_number} \u2014 {batch.section_name}"
            cell = ws.cell(row=row, column=1, value=section_label)
            cell.font = _SECTION_FONT
            cell.fill = _SECTION_FILL
            ws.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=5
            )
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = _SECTION_FILL
                ws.cell(row=row, column=c).border = _BORDER
            row += 1

        if batch.error:
            cell = ws.cell(row=row, column=1, value=batch.page_number)
            ws.cell(row=row, column=2, value="BLAD")
            ws.cell(row=row, column=3, value=_clean(batch.error))
            for c in range(1, 6):
                ws.cell(row=row, column=c).font = _ERROR_FONT
                ws.cell(row=row, column=c).border = _BORDER
            row += 1
            continue

        for tu in batch.units:
            ws.cell(row=row, column=1, value=batch.page_number)
            ws.cell(row=row, column=2, value=f"{tu.category.value}")
            ws.cell(row=row, column=3, value=_clean(tu.source_text))
            ws.cell(row=row, column=4, value=_clean(tu.translated_text))
            ws.cell(row=row, column=5, value=_clean(tu.note_for_designer))

            for c in range(1, 6):
                _apply_cell_style(ws, row, c)
            if tu.note_for_designer:
                _apply_cell_style(ws, row, 5, _NOTE_FONT)

            row += 1

    # Autofilter
    if row > 2:
        ws.auto_filter.ref = f"A1:E{row - 1}"


def _write_glossary_sheet(
    wb: Workbook,
    glossary: GlossaryConfig | None,
) -> None:
    """Glossary reference sheet."""
    ws = wb.create_sheet("Slownik - Glossar")

    headers = ["Termin zrodlowy", f"Tlumaczenie ({glossary.target_lang.upper() if glossary else ''})", "Kontekst"]
    _apply_header_row(ws, 1, headers)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.freeze_panes = "A2"

    if not glossary:
        ws.cell(row=2, column=1, value="Brak slownika")
        return

    row = 2
    for src, tgt in sorted(glossary.terms.items()):
        ws.cell(row=row, column=1, value=src)
        ws.cell(row=row, column=2, value=tgt)
        ws.cell(row=row, column=3, value=glossary.domain)
        for c in range(1, 4):
            _apply_cell_style(ws, row, c)
        row += 1

    if glossary.do_not_translate:
        row += 1
        cell = ws.cell(row=row, column=1, value="NIE TLUMACZ:")
        cell.font = _SECTION_FONT
        cell.fill = _SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
        row += 1
        for term in glossary.do_not_translate:
            ws.cell(row=row, column=1, value=term)
            _apply_cell_style(ws, row, 1)
            row += 1

    if row > 2:
        ws.auto_filter.ref = f"A1:C{row - 1}"


def _write_stats_sheet(
    wb: Workbook,
    batches: list[BatchResult],
    pages: list[list[TranslationUnit]],
    validation: ValidationReport | None,
    target_lang: str,
    source_filename: str,
) -> None:
    """Statistics summary sheet."""
    ws = wb.create_sheet("Statystyki")

    headers = ["Metryka", "Wartosc"]
    _apply_header_row(ws, 1, headers)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30
    ws.freeze_panes = "A2"

    total_units = sum(len(p) for p in pages)
    translated_units = sum(
        len([u for u in b.units if u.translated_text]) for b in batches if not b.error
    )
    total_pages = len(pages)
    source_chars = sum(len(u.source_text) for p in pages for u in p)
    translated_chars = sum(
        len(u.translated_text)
        for b in batches
        if not b.error
        for u in b.units
    )
    failed_batches = sum(1 for b in batches if b.error)

    stats: list[tuple[str, str]] = [
        ("Plik zrodlowy", source_filename),
        ("Liczba stron", str(total_pages)),
        ("Liczba jednostek tlumaczeniowych", str(total_units)),
        ("Przetlumaczonych jednostek", str(translated_units)),
        ("Liczba znakow (oryginal)", f"~{source_chars:,}"),
        ("Liczba znakow (tlumaczenie)", f"~{translated_chars:,}"),
        ("Jezyk docelowy", target_lang.upper()),
        ("Nieudane batche", str(failed_batches)),
        ("Data wygenerowania", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]

    if validation:
        stats.append(("Bledy walidacji", str(validation.errors_count)))
        stats.append(("Ostrzezenia walidacji", str(validation.warnings_count)))

    for i, (metric, value) in enumerate(stats, 2):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value=value)
        for c in (1, 2):
            _apply_cell_style(ws, i, c)


def _write_validation_sheet(
    wb: Workbook,
    validation: ValidationReport | None,
) -> None:
    """Validation issues sheet."""
    ws = wb.create_sheet("Walidacja")

    headers = ["Strona", "ID", "Typ", "Waznosc", "Wiadomosc", "Oryginal", "Tlumaczenie"]
    _apply_header_row(ws, 1, headers)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 35
    ws.freeze_panes = "A2"

    if not validation or not validation.issues:
        ws.cell(row=2, column=1, value="Brak problemow")
        return

    for i, issue in enumerate(validation.issues, 2):
        ws.cell(row=i, column=1, value=issue.page)
        ws.cell(row=i, column=2, value=issue.unit_id)
        ws.cell(row=i, column=3, value=issue.check_type)
        ws.cell(row=i, column=4, value=issue.severity.value)
        ws.cell(row=i, column=5, value=_clean(issue.message))
        ws.cell(row=i, column=6, value=_clean(issue.source_text))
        ws.cell(row=i, column=7, value=_clean(issue.translated_text))

        font = _NORMAL_FONT
        if issue.severity == ValidationSeverity.ERROR:
            font = _ERROR_FONT
        elif issue.severity == ValidationSeverity.WARNING:
            font = _WARNING_FONT

        for c in range(1, 8):
            _apply_cell_style(ws, i, c, font)

    last_row = len(validation.issues) + 1
    ws.auto_filter.ref = f"A1:G{last_row}"


def export_xlsx(
    batches: list[BatchResult],
    pages: list[list[TranslationUnit]],
    validation: ValidationReport | None,
    glossary: GlossaryConfig | None,
    target_lang: str,
    source_filename: str,
) -> bytes:
    """Export full catalog translation to a formatted XLSX file.

    Returns raw bytes suitable for Streamlit download_button.
    """
    wb = Workbook()

    _write_translation_sheet(wb, batches, pages, target_lang)
    _write_glossary_sheet(wb, glossary)
    _write_stats_sheet(wb, batches, pages, validation, target_lang, source_filename)
    _write_validation_sheet(wb, validation)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
